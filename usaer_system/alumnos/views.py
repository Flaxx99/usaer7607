# alumnos/views.py
import os
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from django.conf import settings
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model

from .models import Alumno
from .forms import AlumnoForm

User = get_user_model()

def es_profesor(user):
    """Verifica si el usuario es profesor"""
    return user.is_authenticated and hasattr(user, 'role') and user.role == 'Profesor'

@login_required
@user_passes_test(es_profesor)
def listar_alumnos(request):
    """Lista todos los alumnos asignados al profesor logueado"""
    alumnos = Alumno.objects.filter(profesor=request.user).order_by(
        'apellido_paterno', 'apellido_materno', 'nombres'
    )
    return render(request, 'alumnos/listar.html', {
        'alumnos': alumnos
    })

@login_required
@user_passes_test(es_profesor)
def crear_alumno(request):
    """Permite al profesor crear un nuevo alumno"""
    if request.method == 'POST':
        form = AlumnoForm(request.POST, profesor=request.user)
        if form.is_valid():
            alumno = form.save(commit=False)
            alumno.profesor = request.user
            alumno.escuela = request.user.escuela
            alumno.save()
            messages.success(request, "Alumno registrado correctamente.")
            return redirect('alumnos:listar_alumnos')
    else:
        form = AlumnoForm(profesor=request.user)

    return render(request, 'alumnos/form.html', {
        'form': form,
        'titulo': 'Nuevo Alumno'
    })

@login_required
@user_passes_test(es_profesor)
def editar_alumno(request, pk):
    """Permite al profesor editar los datos de uno de sus alumnos"""
    alumno = get_object_or_404(Alumno, pk=pk, profesor=request.user)
    
    if request.method == 'POST':
        form = AlumnoForm(request.POST, instance=alumno, profesor=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Datos del alumno actualizados.")
            return redirect('alumnos:listar_alumnos')
    else:
        form = AlumnoForm(instance=alumno, profesor=request.user)

    return render(request, 'alumnos/form.html', {
        'form': form,
        'titulo': 'Editar Alumno'
    })

@login_required
@user_passes_test(es_profesor)
def eliminar_alumno(request, pk):
    """Permite al profesor eliminar a uno de sus alumnos con confirmación"""
    alumno = get_object_or_404(Alumno, pk=pk, profesor=request.user)
    
    if request.method == 'POST':
        try:
            nombre_completo = f"{alumno.apellido_paterno} {alumno.apellido_materno}, {alumno.nombres}"
            alumno.delete()
            messages.success(request, f"Alumno {nombre_completo} eliminado correctamente.")
            return redirect('alumnos:listar_alumnos')
        except Exception as e:
            messages.error(request, f"Error al eliminar alumno: {str(e)}")
            return redirect('alumnos:listar_alumnos')

    return render(request, 'alumnos/confirmar_eliminar.html', {
        'alumno': alumno,
        'titulo': 'Confirmar eliminación'
    })

@login_required
@user_passes_test(es_profesor)
def exportar_rac(request):
    """Exporta un archivo Excel con el formato RAC usando plantilla"""
    try:
        plantilla = os.path.join(settings.BASE_DIR, 'static', 'templates', 'template_rac.xlsx')
        wb = openpyxl.load_workbook(plantilla)
        ws = wb.active
        fila = 5  # fila de inicio en la plantilla

        for alumno in Alumno.objects.filter(profesor=request.user):
            ws[f'A{fila}'] = alumno.apellido_paterno or ''
            ws[f'B{fila}'] = alumno.apellido_materno or ''
            ws[f'C{fila}'] = alumno.nombres or ''
            ws[f'D{fila}'] = alumno.curp or ''
            ws[f'E{fila}'] = alumno.sexo or ''
            ws[f'F{fila}'] = alumno.edad or ''
            ws[f'G{fila}'] = alumno.grado or ''
            ws[f'H{fila}'] = alumno.clasificacion or ''
            ws[f'I{fila}'] = alumno.clasificacion_otro or ''
            fila += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=RAC_{request.user.username}.xlsx'
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f"Error al generar el archivo: {str(e)}")
        return redirect('alumnos:listar_alumnos')